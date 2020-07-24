# -*- coding: UTF-8 -*-
import openpyxl as xl
from openpyxl.utils import get_column_letter, column_index_from_string
import json
import os


def readCompose():
    wb = xl.load_workbook('parser_excel/合金表.xlsx')
    sheet = wb["Sheet1"]
    props = [cell.value.replace(' ', '') for cell in sheet[2]]
    map = {}
    for row in sheet.iter_rows(min_row=3):
        no = str(row[1].value).replace(' ', '')
        item = {}
        for prop, cell in zip(props, row):
            item[prop] = str(cell.value).replace(' ', '')
        map[no] = item
    return map


def saveAsJson(obj):
    f = open('parser_excel/合并.txt', 'w', encoding="UTF-8")
    json.dump(obj, f, ensure_ascii=False)
    f.close()

def saveToMain():
    wb = xl.load_workbook('parser_excel/合金表.xlsx')
    sheet = wb["Sheet1"]
    j = sheet.max_column
    for pp_wb in [xl.load_workbook(os.path.join('parser_excel/props', file)) for file in
                  os.listdir('parser_excel/props')]:
        j += 1
        pp_sheet = pp_wb["Sheet1"]
        sheet.cell(2, j).value = pp_sheet["C1"].value
        for row in pp_sheet.iter_rows(min_row=2):
            if row[1].value is None: continue
            no = str(row[1].value).replace(' ', '')
            prop_value = str(row[2].value).replace(' ', '')
            is_exist = False
            for cell in sheet["B"]:
                if str(cell.value).replace(' ', '') == no:
                    i = cell.row
                    is_exist = True
                    break
            if not is_exist:
                print('{}属性表中编号{}在主表中不存在，已跳过本条录入。'.format(pp_sheet["C1"].value,no))
                continue
            if sheet.cell(2, j).value.find('电导率') >= 0:
                prop_value = eval(prop_value.replace('^', '**'))
            if sheet.cell(i, j).value is not None and sheet.cell(i, j).value != '':
                print(pp_sheet["C1"].value + '属性表中，有相同产品编号[' + sheet.cell(i, 2).value + ']冲突,本次属性值不更新，请手动处理')
            else:
                sheet.cell(i, j).value = prop_value

    '''
    临时处理，空值属性设为0
    '''
    for row in sheet.iter_rows(min_row=2,min_col=column_index_from_string('B'),max_col=column_index_from_string('N')):
        if row[0].value is None or row[0].value == '':
            continue
        else:
            for cell in row[1:]:
                if cell.value is None or cell.value == '':
                    cell.value = 0

    wb.save('parser_excel/合金表综合.xlsx')


if __name__ == '__main__':
    saveToMain()
    # combine()
